"""
Question Import/Export functionality for CSV and Excel files
"""
import csv
import io
import json
from typing import List, Dict, Any, Tuple
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import QuestionBank, Question


class QuestionImportExport:
    """Handle import/export of questions from CSV/Excel"""

    # Mapping of respondent types to default categories
    RESPONDENT_TO_CATEGORY_MAPPING = {
        'input_suppliers': 'input_supply',
        'farmers': 'production',
        'aggregators_lbcs': 'distribution',
        'processors': 'processing',
        'processors_eu': 'processing',
        'retailers_food_vendors': 'distribution',
        'retailers_food_vendors_eu': 'distribution',
        'local_consumers': 'consumption',
        'consumers_eu_prolific': 'consumption',
        'client_business_eu_prolific': 'consumption',
        'government': 'governance',
        'ngos': 'governance',
        'certification_schemes': 'certification',
        'coop': 'distribution',
        'chief': 'governance',
    }

    # Template column definitions
    TEMPLATE_COLUMNS = [
        'question_text',
        'response_type',
        'question_category',
        'targeted_respondents',
        'targeted_commodities',
        'targeted_countries',
        'data_source',
        'is_required',
        'options',
        'section_header',
        'section_preamble',
        'is_follow_up',
        'parent_question_text',
        'condition_operator',
        'condition_value',
    ]

    # Column descriptions for template
    COLUMN_DESCRIPTIONS = {
        'question_text': 'The actual question text (REQUIRED)',
        'response_type': 'Type: text_short, text_long, numeric_integer, choice_single, etc. (REQUIRED)',
        'question_category': 'Custom category (e.g., Production, Processing, Finance) - any text (default: general)',
        'targeted_respondents': 'Comma-separated: farmers,processors,retailers_food_vendors (REQUIRED)',
        'targeted_commodities': 'Comma-separated: cocoa,maize,palm_oil (REQUIRED)',
        'targeted_countries': 'Comma-separated: Ghana,Nigeria,Kenya (REQUIRED)',
        'data_source': 'Source: internal, partner_university, partner_ngo, etc. (default: internal)',
        'is_required': 'true or false - Should this question be required? (default: true)',
        'options': 'For choice questions only: Option1|Option2|Option3 (pipe-separated)',
        'section_header': 'Section/group title - questions with same header are grouped together (optional)',
        'section_preamble': 'Introductory text displayed before first question in section (optional)',
        'is_follow_up': 'true or false - Is this a follow-up/conditional question? (default: false)',
        'parent_question_text': 'Text of parent question (for follow-up questions only)',
        'condition_operator': 'Operator: equals, contains, greater_than, in, etc. (for follow-up questions)',
        'condition_value': 'Value(s) to compare against. For "in" operator: Val1|Val2|Val3',
    }

    # Valid choices
    VALID_CATEGORIES = [choice[0] for choice in QuestionBank.CATEGORY_CHOICES]
    VALID_RESPONDENTS = [choice[0] for choice in QuestionBank.RESPONDENT_CHOICES]
    VALID_COMMODITIES = [choice[0] for choice in QuestionBank.COMMODITY_CHOICES]
    VALID_DATA_SOURCES = [choice[0] for choice in QuestionBank.DATA_SOURCE_CHOICES]
    VALID_RESPONSE_TYPES = [
        'text_short', 'text_long', 'numeric_integer', 'numeric_decimal', 'scale_rating',
        'choice_single', 'choice_multiple', 'date', 'datetime', 'geopoint', 'geoshape',
        'image', 'audio', 'video', 'file', 'signature', 'barcode'
    ]

    @classmethod
    def generate_csv_template(cls) -> HttpResponse:
        """Generate CSV template with headers and example row"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="question_template_{timezone.now().strftime("%Y%m%d")}.csv"'

        writer = csv.writer(response)

        # Write headers
        writer.writerow(cls.TEMPLATE_COLUMNS)

        # Write descriptions row
        descriptions = [cls.COLUMN_DESCRIPTIONS.get(col, '') for col in cls.TEMPLATE_COLUMNS]
        writer.writerow(descriptions)

        # Write example row
        example_row = [
            'What is your primary source of income?',
            'choice_single',
            'Production',
            'farmers,aggregators_lbcs',
            'cocoa,maize',
            'Ghana,Nigeria',
            'internal',
            'true',
            'Farming|Trading|Processing|Other',
            '',
            '',
            'false',
            '',
            '',
            '',
        ]
        writer.writerow(example_row)

        # Write a follow-up question example
        followup_row = [
            'How many hectares of farmland do you own?',
            'numeric_decimal',
            'Production',
            'farmers',
            'cocoa,maize',
            'Ghana,Nigeria',
            'internal',
            'true',
            '',
            '',
            '',
            'true',
            'What is your primary source of income?',
            'equals',
            'Farming',
        ]
        writer.writerow(followup_row)

        # Write section with preamble example (first question in section)
        section_example_1 = [
            'How effective do you think better knowledge sharing help solve current challenges in the value chain?',
            'choice_single',
            'Knowledge Sharing',
            'farmers,processors,aggregators_lbcs',
            'cocoa,maize,palm_oil',
            'Ghana,Nigeria',
            'internal',
            'true',
            'Not effective at all|Slightly effective|Somewhat effective|Moderately effective|Effective|Very effective|Extremely effective',
            'Solution 1: Facilitation and Improvement of Knowledge Sharing',
            'This solution is about improving how information, skills, and experiences are exchanged across the value chain.',
            'false',
            '',
            '',
            '',
        ]
        writer.writerow(section_example_1)

        # Write another question in same section (same section_header, empty preamble)
        section_example_2 = [
            'How feasible would it be to implement this solution in your local context?',
            'choice_single',
            'Knowledge Sharing',
            'farmers,processors,aggregators_lbcs',
            'cocoa,maize,palm_oil',
            'Ghana,Nigeria',
            'internal',
            'true',
            'Not feasible at all|Slightly feasible|Somewhat feasible|Moderately feasible|Feasible|Very feasible|Extremely feasible',
            'Solution 1: Facilitation and Improvement of Knowledge Sharing',
            '',
            'false',
            '',
            '',
            '',
        ]
        writer.writerow(section_example_2)

        return response

    @classmethod
    def generate_excel_template(cls) -> HttpResponse:
        """Generate Excel template with formatting and validation"""
        workbook = openpyxl.Workbook()

        # Create main sheet
        sheet = workbook.active
        sheet.title = "Questions"

        # Style definitions
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        desc_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Write headers
        for col_idx, column in enumerate(cls.TEMPLATE_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write descriptions
        for col_idx, column in enumerate(cls.TEMPLATE_COLUMNS, start=1):
            cell = sheet.cell(row=2, column=col_idx, value=cls.COLUMN_DESCRIPTIONS.get(column, ''))
            cell.fill = desc_fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Write example row (regular question)
        example_row = [
            'What is your primary source of income?',
            'choice_single',
            'Production',
            'farmers,aggregators_lbcs',
            'cocoa,maize',
            'Ghana,Nigeria',
            'internal',
            'true',
            'Farming|Trading|Processing|Other',
            '',
            '',
            'false',
            '',
            '',
            '',
        ]
        for col_idx, value in enumerate(example_row, start=1):
            cell = sheet.cell(row=3, column=col_idx, value=value)
            cell.fill = example_fill

        # Write follow-up question example
        followup_row = [
            'How many hectares of farmland do you own?',
            'numeric_decimal',
            'Production',
            'farmers',
            'cocoa,maize',
            'Ghana,Nigeria',
            'internal',
            'true',
            '',
            '',
            '',
            'true',
            'What is your primary source of income?',
            'equals',
            'Farming',
        ]
        for col_idx, value in enumerate(followup_row, start=1):
            cell = sheet.cell(row=4, column=col_idx, value=value)
            cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")

        # Write section with preamble example (first question in section)
        section_fill = PatternFill(start_color="E7E6FF", end_color="E7E6FF", fill_type="solid")
        section_row_1 = [
            'How effective do you think better knowledge sharing help solve current challenges in the value chain?',
            'choice_single',
            'Knowledge Sharing',
            'farmers,processors,aggregators_lbcs',
            'cocoa,maize,palm_oil',
            'Ghana,Nigeria',
            'internal',
            'true',
            'Not effective at all|Slightly effective|Somewhat effective|Moderately effective|Effective|Very effective|Extremely effective',
            'Solution 1: Facilitation and Improvement of Knowledge Sharing',
            'This solution is about improving how information, skills, and experiences are exchanged across the value chain.',
            'false',
            '',
            '',
            '',
        ]
        for col_idx, value in enumerate(section_row_1, start=1):
            cell = sheet.cell(row=5, column=col_idx, value=value)
            cell.fill = section_fill

        # Write another question in same section (same section_header, empty preamble)
        section_row_2 = [
            'How feasible would it be to implement this solution in your local context?',
            'choice_single',
            'Knowledge Sharing',
            'farmers,processors,aggregators_lbcs',
            'cocoa,maize,palm_oil',
            'Ghana,Nigeria',
            'internal',
            'true',
            'Not feasible at all|Slightly feasible|Somewhat feasible|Moderately feasible|Feasible|Very feasible|Extremely feasible',
            'Solution 1: Facilitation and Improvement of Knowledge Sharing',
            '',
            'false',
            '',
            '',
            '',
        ]
        for col_idx, value in enumerate(section_row_2, start=1):
            cell = sheet.cell(row=6, column=col_idx, value=value)
            cell.fill = section_fill

        # Adjust column widths (updated for new columns)
        column_widths = [50, 20, 20, 35, 30, 25, 25, 15, 40, 40, 50, 15, 40, 20, 30]
        for col_idx, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Set row heights
        sheet.row_dimensions[1].height = 20
        sheet.row_dimensions[2].height = 40

        # Create reference sheet with valid values
        ref_sheet = workbook.create_sheet("Reference Values")
        ref_sheet.sheet_state = 'hidden'  # Hide by default

        # Add valid values
        ref_data = {
            'A': ('Categories', cls.VALID_CATEGORIES),
            'B': ('Response Types', cls.VALID_RESPONSE_TYPES),
            'C': ('Respondents', cls.VALID_RESPONDENTS),
            'D': ('Commodities', cls.VALID_COMMODITIES),
            'E': ('Data Sources', cls.VALID_DATA_SOURCES),
        }

        for col, (title, values) in ref_data.items():
            ref_sheet[f'{col}1'] = title
            ref_sheet[f'{col}1'].font = Font(bold=True)
            for idx, value in enumerate(values, start=2):
                ref_sheet[f'{col}{idx}'] = value

        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="question_template_{timezone.now().strftime("%Y%m%d")}.xlsx"'

        return response

    @classmethod
    def parse_csv(cls, file) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse CSV file and return list of question data dictionaries and errors"""
        questions = []
        errors = []

        try:
            # Read file content
            content = file.read().decode('utf-8-sig')  # Handle BOM
            reader = csv.DictReader(io.StringIO(content))

            # Validate headers
            required_cols = ['question_text', 'response_type', 'targeted_respondents', 'targeted_commodities', 'targeted_countries']
            if not all(col in reader.fieldnames for col in required_cols):
                errors.append(f"CSV must contain all required columns: {', '.join(required_cols)}")
                return [], errors

            for row_num, row in enumerate(reader, start=2):  # Start at 2 (header + data rows)
                # Skip description row (row 2 contains column descriptions)
                if row_num == 2 and any(desc in str(row.get('question_text', '')) for desc in ['The actual question', 'REQUIRED']):
                    continue

                # Skip empty rows
                if not row.get('question_text', '').strip():
                    continue

                question_data, row_errors = cls._parse_row(row, row_num)

                if row_errors:
                    errors.extend([f"Row {row_num}: {err}" for err in row_errors])
                else:
                    questions.append(question_data)

        except Exception as e:
            errors.append(f"Failed to parse CSV: {str(e)}")

        return questions, errors

    @classmethod
    def parse_excel(cls, file) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse Excel file and return list of question data dictionaries and errors"""
        questions = []
        errors = []

        try:
            workbook = openpyxl.load_workbook(file, read_only=True)
            sheet = workbook.active

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(cell.value)

            # Validate headers
            required_cols = ['question_text', 'response_type', 'targeted_respondents', 'targeted_commodities', 'targeted_countries']
            if not all(col in headers for col in required_cols):
                errors.append(f"Excel must contain all required columns: {', '.join(required_cols)}")
                return [], errors

            # Parse data rows (skip header, description, and example rows)
            for row_num, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
                # Skip empty rows
                if not row[0] or not str(row[0]).strip():
                    continue

                # Create row dictionary
                row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}

                question_data, row_errors = cls._parse_row(row_dict, row_num)

                if row_errors:
                    errors.extend([f"Row {row_num}: {err}" for err in row_errors])
                else:
                    questions.append(question_data)

            workbook.close()

        except Exception as e:
            errors.append(f"Failed to parse Excel: {str(e)}")

        return questions, errors

    @classmethod
    def _parse_row(cls, row: Dict[str, Any], row_num: int, project=None) -> Tuple[Dict[str, Any], List[str]]:
        """
        Parse a single row and return question data and errors.

        Args:
            row: Dictionary containing question data
            row_num: Row number for error reporting
            project: Optional Project instance to validate partner names
        """
        errors = []

        # Required fields
        question_text = str(row.get('question_text', '')).strip()
        response_type = str(row.get('response_type', '')).strip().lower()

        if not question_text:
            errors.append("question_text is required")

        if not response_type:
            errors.append("response_type is required")
        elif response_type not in cls.VALID_RESPONSE_TYPES:
            errors.append(f"Invalid response_type '{response_type}'. Valid: {', '.join(cls.VALID_RESPONSE_TYPES)}")

        # Parse data source (set defaults for removed fields)
        data_source = str(row.get('data_source', 'internal')).strip().lower()
        if data_source and data_source not in cls.VALID_DATA_SOURCES:
            errors.append(f"Invalid data_source '{data_source}'. Valid: {', '.join(cls.VALID_DATA_SOURCES)}")
        if not data_source:
            data_source = 'internal'

        # Set default values for removed fields
        research_partner_name = ''
        research_partner_contact = ''
        work_package = ''
        priority_score = 5

        # Parse boolean fields
        is_required = cls._parse_boolean(row.get('is_required'), default=True)
        is_follow_up = cls._parse_boolean(row.get('is_follow_up'), default=False)

        # Parse options (pipe-separated)
        options = []
        options_str = str(row.get('options', '')).strip()
        if options_str and options_str != '{}':
            options = [opt.strip() for opt in options_str.split('|') if opt.strip()]

        # Validate options for choice types
        if response_type in ['choice_single', 'choice_multiple'] and not options:
            errors.append(f"options required for response_type '{response_type}'")

        # Parse conditional logic fields
        parent_question_text = str(row.get('parent_question_text', '')).strip()
        condition_operator = str(row.get('condition_operator', '')).strip().lower()
        condition_value_str = str(row.get('condition_value', '')).strip()

        conditional_logic = None
        if is_follow_up:
            # Validate follow-up question fields
            if not parent_question_text:
                errors.append("parent_question_text required for follow-up questions")
            if not condition_operator:
                errors.append("condition_operator required for follow-up questions")
            else:
                # Validate operator
                valid_operators = ['equals', 'not_equals', 'contains', 'not_contains',
                                  'greater_than', 'less_than', 'greater_or_equal',
                                  'less_or_equal', 'in', 'not_in', 'is_empty',
                                  'is_not_empty', 'between']
                if condition_operator not in valid_operators:
                    errors.append(f"Invalid condition_operator '{condition_operator}'. Valid: {', '.join(valid_operators)}")

            # Build conditional logic (parent_question_id will be resolved during import)
            if not errors:
                condition_value = None
                condition_values = None

                # Parse value(s) based on operator
                if condition_operator in ['in', 'not_in', 'between']:
                    # Multiple values (pipe-separated)
                    condition_values = [v.strip() for v in condition_value_str.split('|') if v.strip()]
                    if not condition_values:
                        errors.append(f"condition_value required for operator '{condition_operator}'")
                elif condition_operator not in ['is_empty', 'is_not_empty']:
                    # Single value
                    condition_value = condition_value_str
                    if not condition_value:
                        errors.append(f"condition_value required for operator '{condition_operator}'")

                # Build conditional logic structure (parent_question_id placeholder)
                conditional_logic = {
                    'enabled': True,
                    'parent_question_text': parent_question_text,  # Store text for now, will resolve to ID during import
                    'show_if': {
                        'operator': condition_operator,
                    }
                }

                if condition_value is not None:
                    conditional_logic['show_if']['value'] = condition_value
                if condition_values is not None:
                    conditional_logic['show_if']['values'] = condition_values

        # Parse list fields (comma-separated)
        targeted_respondents = cls._parse_list_field(row.get('targeted_respondents'))
        if not targeted_respondents:
            errors.append("targeted_respondents is required")
        for respondent in targeted_respondents:
            if respondent not in cls.VALID_RESPONDENTS:
                errors.append(f"Invalid respondent '{respondent}'. Valid: {', '.join(cls.VALID_RESPONDENTS)}")

        # Parse question_category from the row (user-specified custom category)
        question_category = str(row.get('question_category', 'general')).strip()
        if not question_category:
            question_category = 'general'

        targeted_commodities = cls._parse_list_field(row.get('targeted_commodities'))
        if not targeted_commodities:
            errors.append("targeted_commodities is required")
        for commodity in targeted_commodities:
            if commodity not in cls.VALID_COMMODITIES:
                errors.append(f"Invalid commodity '{commodity}'. Valid: {', '.join(cls.VALID_COMMODITIES)}")

        targeted_countries = cls._parse_list_field(row.get('targeted_countries'))
        if not targeted_countries:
            errors.append("targeted_countries is required")

        # Determine question ownership and sources based on data_source
        is_owner_question = True
        question_sources = ['owner']
        
        if data_source == 'internal':
            # Internal questions belong to owner only
            is_owner_question = True
            question_sources = ['owner']
        elif data_source == 'collaborative':
            # Collaborative questions belong to both owner and partner
            is_owner_question = True
            if research_partner_name:
                question_sources = ['owner', research_partner_name]
            else:
                question_sources = ['owner']
                errors.append("Collaborative questions require research_partner_name")
        elif research_partner_name:
            # Partner questions belong to the partner only
            is_owner_question = False
            question_sources = [research_partner_name]
        else:
            # Fallback: if no partner name specified, treat as owner
            is_owner_question = True
            question_sources = ['owner']

        # Parse section fields
        section_header = str(row.get('section_header', '')).strip()
        section_preamble = str(row.get('section_preamble', '')).strip()

        # Build question data
        question_data = {
            'question_text': question_text,
            'question_category': question_category,
            'response_type': response_type,
            'is_required': is_required,
            'allow_multiple': False,
            'options': options if options else None,
            'validation_rules': None,
            'targeted_respondents': targeted_respondents,
            'targeted_commodities': targeted_commodities,
            'targeted_countries': targeted_countries,
            'data_source': data_source,
            'research_partner_name': research_partner_name,
            'research_partner_contact': research_partner_contact,
            'work_package': work_package,
            'priority_score': priority_score,
            'is_owner_question': is_owner_question,
            'question_sources': question_sources,
            'tags': [],
            'is_active': True,
            'is_follow_up': is_follow_up,
            'conditional_logic': conditional_logic,
            'section_header': section_header,
            'section_preamble': section_preamble,
        }

        return question_data, errors

    @staticmethod
    def _parse_boolean(value: Any, default: bool = False) -> bool:
        """Parse boolean value from various formats"""
        if value is None or value == '':
            return default
        if isinstance(value, bool):
            return value
        str_value = str(value).strip().lower()
        return str_value in ['true', '1', 'yes', 'y']

    @staticmethod
    def _parse_list_field(value: Any) -> List[str]:
        """Parse comma-separated list field"""
        if not value or str(value).strip() == '':
            return []
        return [item.strip() for item in str(value).split(',') if item.strip()]

    @classmethod
    def import_questions_to_bank(cls, questions_data: List[Dict[str, Any]], project, created_by_user, created_by: str = '') -> Dict[str, Any]:
        """
        Import questions to QuestionBank for a specific project.

        Args:
            questions_data: List of question data dictionaries
            project: Project instance (required - question banks are project-specific)
            created_by_user: User instance who is creating these questions
            created_by: Legacy string field for backward compatibility
        """
        created_count = 0
        updated_count = 0
        errors = []
        question_text_to_id = {}  # Map question text to QuestionBank ID

        print(f"\n=== IMPORT DEBUG ===")
        print(f"Total questions to import: {len(questions_data)}")
        for i, qd in enumerate(questions_data):
            print(f"  {i+1}. '{qd.get('question_text')[:60]}...' (is_follow_up={qd.get('is_follow_up')})")
        print()

        # First pass: Import/update regular questions
        print("FIRST PASS: Regular questions")
        for question_data in questions_data:
            # Skip follow-up questions in first pass
            if question_data.get('is_follow_up'):
                continue

            try:
                # Check if question already exists
                # A duplicate is: same question_text + same respondent type list + same section_preamble + same project
                # This allows the same question for different respondent types or different sections
                targeted_respondents = question_data.get('targeted_respondents', [])
                section_preamble = question_data.get('section_preamble', '') or ''

                # Build query filter for duplicate detection
                query_filter = {
                    'question_text': question_data['question_text'],
                    'project': project,
                    'section_preamble': section_preamble
                }

                # Filter by exact match of targeted_respondents list
                existing = None
                candidates = QuestionBank.objects.filter(**query_filter)
                for candidate in candidates:
                    # Compare the full targeted_respondents list (order matters)
                    if candidate.targeted_respondents == targeted_respondents:
                        existing = candidate
                        break

                if existing:
                    # Update existing question
                    for key, value in question_data.items():
                        if key not in ['project', 'created_by_user']:  # Don't update these
                            setattr(existing, key, value)
                    existing.save()
                    updated_count += 1
                    question_text_to_id[question_data['question_text']] = existing.id
                else:
                    # Create new question
                    question_data['created_by'] = created_by
                    question_data['project'] = project
                    question_data['created_by_user'] = created_by_user
                    new_question = QuestionBank.objects.create(**question_data)
                    created_count += 1
                    question_text_to_id[question_data['question_text']] = new_question.id

            except Exception as e:
                errors.append(f"Failed to import '{question_data.get('question_text', 'Unknown')}': {str(e)}")

        print(f"\nFirst pass complete. Created mapping for {len(question_text_to_id)} questions:")
        for q_text, q_id in question_text_to_id.items():
            print(f"  '{q_text}' -> {q_id}")
        print()

        # Second pass: Import/update follow-up questions (resolve parent references)
        print("SECOND PASS: Follow-up questions")
        for question_data in questions_data:
            # Only process follow-up questions
            if not question_data.get('is_follow_up'):
                continue

            try:
                # Resolve parent question ID from text
                conditional_logic = question_data.get('conditional_logic')
                if conditional_logic:
                    parent_text = conditional_logic.get('parent_question_text')
                    print(f"  Processing follow-up: '{question_data.get('question_text')[:50]}'")
                    print(f"    Looking for parent: '{parent_text}'")
                    print(f"    Parent text repr: {repr(parent_text)}")
                    print(f"    Available keys in mapping:")
                    for key in question_text_to_id.keys():
                        print(f"      - {repr(key)}")
                    print(f"    Parent in mapping? {parent_text in question_text_to_id}")
                    if parent_text in question_text_to_id:
                        # Convert parent_question_text to parent_question_id
                        conditional_logic['parent_question_id'] = str(question_text_to_id[parent_text])
                        del conditional_logic['parent_question_text']
                    else:
                        # Try to find existing parent question in the same project
                        parent_query = {
                            'question_text': parent_text,
                            'project': project
                        }
                        parent_question = QuestionBank.objects.filter(**parent_query).first()
                        if parent_question:
                            conditional_logic['parent_question_id'] = str(parent_question.id)
                            del conditional_logic['parent_question_text']
                        else:
                            errors.append(f"Parent question '{parent_text}' not found for follow-up question '{question_data.get('question_text')}'")
                            continue

                # Check if question already exists (same duplicate logic as first pass)
                targeted_respondents = question_data.get('targeted_respondents', [])
                section_preamble = question_data.get('section_preamble', '') or ''

                query_filter = {
                    'question_text': question_data['question_text'],
                    'project': project,
                    'section_preamble': section_preamble
                }

                # Filter by exact match of targeted_respondents list
                existing = None
                candidates = QuestionBank.objects.filter(**query_filter)
                for candidate in candidates:
                    # Compare the full targeted_respondents list (order matters)
                    if candidate.targeted_respondents == targeted_respondents:
                        existing = candidate
                        break

                if existing:
                    # Update existing question
                    for key, value in question_data.items():
                        if key not in ['project', 'created_by_user']:  # Don't update these
                            setattr(existing, key, value)
                    existing.save()
                    updated_count += 1
                else:
                    # Create new question
                    question_data['created_by'] = created_by
                    question_data['project'] = project
                    question_data['created_by_user'] = created_by_user
                    QuestionBank.objects.create(**question_data)
                    created_count += 1

            except Exception as e:
                errors.append(f"Failed to import follow-up '{question_data.get('question_text', 'Unknown')}': {str(e)}")

        return {
            'created': created_count,
            'updated': updated_count,
            'errors': errors,
            'total_processed': created_count + updated_count,
        }
